from __future__ import annotations

import hashlib
import json
import re
import zipfile
from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import TypeAlias

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

CellValue: TypeAlias = str | int | bool | date | datetime | Decimal | None
Row: TypeAlias = Mapping[str, CellValue]

DATASET_VERSION = "1.0.0"
SCHEMA_VERSION = "1.0"
GENERATOR_SEED = 20260731
GENERATED_AT = "2026-07-31T12:00:00+03:00"
DATE_RANGE_START = date(2026, 2, 1)
DATE_RANGE_END = date(2026, 7, 31)
MONTHS = tuple(range(2, 8))
SITE_SLUGS = (
    "ulubatli-sitesi",
    "cinarpark-apartmani",
    "mavisehir-konutlari",
    "ihlamur-residence",
    "gunes-evleri",
)
MONEY_COLUMNS = {"amount"}
DATE_COLUMNS = {
    "charge_date",
    "due_date",
    "payment_date",
    "expense_date",
    "valid_until",
}
DATETIME_COLUMNS = {"published_at"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonicalize_xlsx(path: Path) -> None:
    temporary = path.with_suffix(".canonical.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temporary,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, date_time=(2026, 7, 31, 12, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            content = source.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    (
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                        b"2026-07-31T09:00:00Z</dcterms:modified>"
                    ),
                    content,
                )
            target.writestr(info, content)
    temporary.replace(path)


def write_workbook(
    path: Path,
    *,
    sheet_name: str,
    columns: Sequence[str],
    rows: Iterable[Row],
) -> int:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(list(columns))
    row_count = 0
    for row in rows:
        worksheet.append([row.get(column) for column in columns])
        row_count += 1

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, column in enumerate(columns, start=1):
        width = max(
            len(column),
            max(
                (
                    len(str(worksheet.cell(row=row, column=index).value or ""))
                    for row in range(2, worksheet.max_row + 1)
                ),
                default=0,
            ),
        )
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(width + 2, 12),
            42,
        )
        if column in DATE_COLUMNS:
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = "yyyy-mm-dd"
        elif column in DATETIME_COLUMNS:
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        elif column in MONEY_COLUMNS:
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = '#,##0.00'

    fixed_time = datetime(2026, 7, 31, 9, 0, 0)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    canonicalize_xlsx(path)
    return row_count


def read_workbook(path: Path) -> list[dict[str, CellValue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(f"{path}: workbook tek sheet içermelidir.")
    worksheet = workbook[workbook.sheetnames[0]]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        raise ValueError(f"{path}: başlık satırı bulunamadı.")
    headers = [str(value) for value in values[0]]
    return [dict(zip(headers, row, strict=True)) for row in values[1:]]


def write_json(path: Path, value: object) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

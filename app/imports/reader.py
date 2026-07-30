from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, cast
from zoneinfo import ZoneInfo

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.imports.constants import SUPPORTED_CURRENCY, SUPPORTED_SCHEMA_VERSION
from app.imports.exceptions import PackageValidationError
from app.imports.schemas import (
    BankTransactionRow,
    ChargeRow,
    DemoAnnouncementRow,
    ExpenseRow,
    PaymentRow,
    ResidentUnitRow,
    SiteRow,
    StandardPackage,
)

SITE_HEADERS = (
    "source_site_key",
    "site_name",
    "site_slug",
    "city",
    "district",
    "address_line",
    "currency",
    "is_active",
)
UNIT_HEADERS = (
    "source_unit_key",
    "block_name",
    "unit_number",
    "floor_label",
    "resident_source_key",
    "resident_full_name",
    "phone",
    "email",
    "access_role",
    "is_active",
)
CHARGE_HEADERS = (
    "source_charge_key",
    "source_unit_key",
    "charge_date",
    "due_date",
    "period_year",
    "period_month",
    "charge_type",
    "title",
    "description",
    "amount",
    "currency",
    "status",
)
PAYMENT_HEADERS = (
    "source_payment_key",
    "source_unit_key",
    "payment_date",
    "payment_method",
    "amount",
    "currency",
    "reference",
    "description",
    "status",
)
EXPENSE_HEADERS = (
    "source_expense_key",
    "expense_date",
    "category",
    "vendor_name",
    "description",
    "amount",
    "currency",
    "document_number",
    "status",
)
ANNOUNCEMENT_HEADERS = (
    "source_announcement_key",
    "published_at",
    "title",
    "body",
    "priority",
    "valid_until",
    "is_active",
)
CHARGE_TYPES = {"monthly_due", "additional_due", "manual"}
PAYMENT_METHODS = {"cash", "bank_transfer", "card", "other"}
EXPENSE_CATEGORIES = {
    "cleaning",
    "electricity",
    "water",
    "elevator_maintenance",
    "security",
    "landscaping",
    "management_service",
    "technical_maintenance",
    "insurance",
    "repair",
}
ANNOUNCEMENT_PRIORITIES = {"normal", "important", "urgent"}
DEMO_WORKBOOK_NAMES = {
    "B001_Yapibina_Cinar_Apartmani_Degisken_Aidat_Demo.xlsx",
    "B002_Mavişehir_Apartmanı_Degisken_Aidat_Demo.xlsx",
    "B003_Ihlamur_Apartmanı_Degisken_Aidat_Demo.xlsx",
    "B004_Güneş_Apartmanı_Degisken_Aidat_Demo.xlsx",
    "B005_Deniz_Apartmanı_Degisken_Aidat_Demo.xlsx",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _payload_hash(row: dict[str, object]) -> str:
    normalized = {
        key: (
            value.isoformat()
            if isinstance(value, (date, datetime))
            else str(value)
            if isinstance(value, Decimal)
            else value
        )
        for key, value in sorted(row.items())
    }
    encoded = json.dumps(
        normalized,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


def _read_rows(path: Path, expected_headers: tuple[str, ...]) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as error:
        raise PackageValidationError(f"Excel dosyası okunamadı: {path}") from error
    if len(workbook.sheetnames) != 1:
        raise PackageValidationError(f"Excel tek sheet içermelidir: {path}")
    values = list(workbook.active.iter_rows(values_only=True))
    if not values:
        raise PackageValidationError(f"Excel başlık satırı içermiyor: {path}")
    headers = tuple(str(value) for value in values[0])
    if headers != expected_headers:
        raise PackageValidationError(f"Excel kolonları sözleşmeyle eşleşmiyor: {path}")
    return [
        dict(zip(headers, cast(tuple[object, ...], values_row), strict=True))
        for values_row in values[1:]
    ]


def _required_text(row: dict[str, object], field: str) -> str:
    value = row.get(field)
    if not isinstance(value, str) or not value.strip():
        raise PackageValidationError(f"{field}: zorunlu metin bekleniyor.")
    return value.strip()


def _optional_text(row: dict[str, object], field: str) -> str | None:
    value = row.get(field)
    if value is None:
        return None
    if not isinstance(value, str):
        raise PackageValidationError(f"{field}: metin bekleniyor.")
    return value.strip() or None


def _date_value(row: dict[str, object], field: str) -> date:
    value = row.get(field)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError as error:
            raise PackageValidationError(f"{field}: geçersiz tarih.") from error
    raise PackageValidationError(f"{field}: tarih bekleniyor.")


def _integer(row: dict[str, object], field: str) -> int:
    value = row.get(field)
    if isinstance(value, bool):
        raise PackageValidationError(f"{field}: integer bekleniyor.")
    try:
        return int(cast(Any, value))
    except (TypeError, ValueError) as error:
        raise PackageValidationError(f"{field}: integer bekleniyor.") from error


def _decimal(row: dict[str, object], field: str) -> Decimal:
    value = row.get(field)
    if isinstance(value, (bool, str)) or value is None:
        raise PackageValidationError(f"{field}: numeric tutar bekleniyor.")
    try:
        result = Decimal(str(value))
    except InvalidOperation as error:
        raise PackageValidationError(f"{field}: geçersiz tutar.") from error
    if result <= 0 or result.quantize(Decimal("0.01")) != result:
        raise PackageValidationError(f"{field}: pozitif iki ondalıklı tutar gerekli.")
    return result.quantize(Decimal("0.01"))


def _boolean(row: dict[str, object], field: str) -> bool:
    value = row.get(field)
    if not isinstance(value, bool):
        raise PackageValidationError(f"{field}: boolean bekleniyor.")
    return value


def _currency_and_status(row: dict[str, object]) -> None:
    if row.get("currency") != SUPPORTED_CURRENCY:
        raise PackageValidationError("Yalnız TRY para birimi desteklenir.")
    if row.get("status") != "posted":
        raise PackageValidationError("Yalnız posted finans kayıtları desteklenir.")


def _site_rows(rows: list[dict[str, object]]) -> tuple[SiteRow, ...]:
    result: list[SiteRow] = []
    for row in rows:
        if row.get("currency") != SUPPORTED_CURRENCY:
            raise PackageValidationError("Site para birimi TRY olmalıdır.")
        result.append(
            SiteRow(
                source_site_key=_required_text(row, "source_site_key"),
                site_name=_required_text(row, "site_name"),
                site_slug=_required_text(row, "site_slug"),
                city=_required_text(row, "city"),
                district=_required_text(row, "district"),
                address_line=_required_text(row, "address_line"),
                is_active=_boolean(row, "is_active"),
                payload_hash=_payload_hash(row),
            )
        )
    return tuple(result)


def _unit_rows(
    source_site_key: str,
    rows: list[dict[str, object]],
) -> tuple[ResidentUnitRow, ...]:
    result: list[ResidentUnitRow] = []
    for row in rows:
        if row.get("access_role") != "resident":
            raise PackageValidationError("Yalnız resident access_role desteklenir.")
        email = _required_text(row, "email").lower()
        if "@" not in email:
            raise PackageValidationError("Resident e-posta değeri geçersiz.")
        result.append(
            ResidentUnitRow(
                source_site_key=source_site_key,
                source_unit_key=_required_text(row, "source_unit_key"),
                block_name=_optional_text(row, "block_name"),
                unit_number=_required_text(row, "unit_number"),
                floor_label=_optional_text(row, "floor_label"),
                resident_source_key=_required_text(row, "resident_source_key"),
                resident_full_name=_required_text(row, "resident_full_name"),
                phone=_optional_text(row, "phone"),
                email=email,
                is_active=_boolean(row, "is_active"),
                initial_password=None,
                payload_hash=_payload_hash(row),
            )
        )
    return tuple(result)


def _charge_rows(
    source_site_key: str,
    rows: list[dict[str, object]],
) -> tuple[ChargeRow, ...]:
    result: list[ChargeRow] = []
    for row in rows:
        _currency_and_status(row)
        charge_type = _required_text(row, "charge_type")
        if charge_type not in CHARGE_TYPES:
            raise PackageValidationError("Desteklenmeyen charge_type.")
        month = _integer(row, "period_month")
        if not 1 <= month <= 12:
            raise PackageValidationError("period_month 1-12 arasında olmalıdır.")
        result.append(
            ChargeRow(
                source_site_key=source_site_key,
                source_charge_key=_required_text(row, "source_charge_key"),
                source_unit_key=_required_text(row, "source_unit_key"),
                charge_date=_date_value(row, "charge_date"),
                due_date=_date_value(row, "due_date"),
                period_year=_integer(row, "period_year"),
                period_month=month,
                charge_type=charge_type,
                title=_required_text(row, "title"),
                description=_optional_text(row, "description"),
                amount=_decimal(row, "amount"),
                payload_hash=_payload_hash(row),
            )
        )
    return tuple(result)


def _payment_rows(
    source_site_key: str,
    rows: list[dict[str, object]],
) -> tuple[PaymentRow, ...]:
    result: list[PaymentRow] = []
    for row in rows:
        _currency_and_status(row)
        method = _required_text(row, "payment_method")
        if method not in PAYMENT_METHODS:
            raise PackageValidationError("Desteklenmeyen payment_method.")
        result.append(
            PaymentRow(
                source_site_key=source_site_key,
                source_payment_key=_required_text(row, "source_payment_key"),
                source_unit_key=_required_text(row, "source_unit_key"),
                payment_date=_date_value(row, "payment_date"),
                payment_method=method,
                amount=_decimal(row, "amount"),
                reference=_optional_text(row, "reference"),
                description=_optional_text(row, "description"),
                target_charge_source_key=None,
                payload_hash=_payload_hash(row),
            )
        )
    return tuple(result)


def _validate_deferred_rows(
    path: Path,
    *,
    headers: tuple[str, ...],
    kind: str,
) -> int:
    rows = _read_rows(path, headers)
    for row in rows:
        if kind == "expense":
            _currency_and_status(row)
            if row.get("category") not in EXPENSE_CATEGORIES:
                raise PackageValidationError("Desteklenmeyen expense category.")
            _date_value(row, "expense_date")
            _decimal(row, "amount")
        else:
            if row.get("priority") not in ANNOUNCEMENT_PRIORITIES:
                raise PackageValidationError("Desteklenmeyen announcement priority.")
            published = _required_text(row, "published_at")
            try:
                if datetime.fromisoformat(published).utcoffset() is None:
                    raise ValueError
            except ValueError as error:
                raise PackageValidationError(
                    "published_at timezone-aware ISO değer olmalıdır."
                ) from error
            _boolean(row, "is_active")
    return len(rows)


def read_standard_package(path: Path) -> StandardPackage:
    root = path.resolve()
    manifest_path = root / "manifest.json"
    if not manifest_path.is_file():
        return _read_demo_workbooks(root)
    try:
        manifest = cast(
            dict[str, Any],
            json.loads(manifest_path.read_text(encoding="utf-8")),
        )
    except (OSError, json.JSONDecodeError) as error:
        raise PackageValidationError("manifest.json okunamadı.") from error
    if manifest.get("schema_version") != SUPPORTED_SCHEMA_VERSION:
        raise PackageValidationError("Desteklenmeyen schema_version.")
    files = cast(list[dict[str, Any]], manifest.get("files"))
    if not isinstance(files, list):
        raise PackageValidationError("Manifest files listesi geçersiz.")
    declared: dict[str, str] = {}
    for item in files:
        relative = str(item.get("relative_path", ""))
        expected_hash = str(item.get("sha256", ""))
        file_path = root / relative
        if not relative or not file_path.is_file():
            raise PackageValidationError(f"Manifest dosyası bulunamadı: {relative}")
        actual_hash = _sha256(file_path)
        if actual_hash != expected_hash:
            raise PackageValidationError(f"Manifest hash uyuşmuyor: {relative}")
        declared[relative] = expected_hash

    sites = _site_rows(_read_rows(root / "sites.xlsx", SITE_HEADERS))
    expected_paths = {"sites.xlsx"}
    units: list[ResidentUnitRow] = []
    charges: list[ChargeRow] = []
    payments: list[PaymentRow] = []
    expense_count = 0
    announcement_count = 0
    for site in sites:
        prefix = site.site_slug
        paths = {
            "units": f"{prefix}/residents_and_units.xlsx",
            "charges": f"{prefix}/charges.xlsx",
            "payments": f"{prefix}/payments.xlsx",
            "expenses": f"{prefix}/expenses.xlsx",
            "announcements": f"{prefix}/announcements.xlsx",
        }
        expected_paths.update(paths.values())
        units.extend(
            _unit_rows(
                site.source_site_key,
                _read_rows(root / paths["units"], UNIT_HEADERS),
            )
        )
        charges.extend(
            _charge_rows(
                site.source_site_key,
                _read_rows(root / paths["charges"], CHARGE_HEADERS),
            )
        )
        payments.extend(
            _payment_rows(
                site.source_site_key,
                _read_rows(root / paths["payments"], PAYMENT_HEADERS),
            )
        )
        expense_count += _validate_deferred_rows(
            root / paths["expenses"],
            headers=EXPENSE_HEADERS,
            kind="expense",
        )
        announcement_count += _validate_deferred_rows(
            root / paths["announcements"],
            headers=ANNOUNCEMENT_HEADERS,
            kind="announcement",
        )
    if set(declared) != expected_paths:
        raise PackageValidationError("Manifest beklenen standart dosya setiyle eşleşmiyor.")

    fingerprint_material = json.dumps(
        {
            "schema_version": manifest["schema_version"],
            "files": sorted(declared.items()),
        },
        separators=(",", ":"),
        sort_keys=True,
    ).encode()
    return StandardPackage(
        root=root,
        dataset_name=str(manifest.get("dataset_name", "")),
        dataset_version=str(manifest.get("dataset_version", "")),
        schema_version=str(manifest["schema_version"]),
        manifest_sha256=_sha256(manifest_path),
        fingerprint=hashlib.sha256(fingerprint_material).hexdigest(),
        sites=sites,
        units=tuple(units),
        charges=tuple(charges),
        payments=tuple(payments),
        expense_count=expense_count,
        announcement_count=announcement_count,
    )


def _money(value: object, *, allow_zero: bool = False) -> Decimal:
    try:
        result = Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError) as error:
        raise PackageValidationError("Demo Excel tutarı geçersiz.") from error
    if result < 0 or (not allow_zero and result == 0):
        raise PackageValidationError("Demo Excel tutarı negatif veya sıfır olamaz.")
    return result


def _demo_rows(workbook: Any, sheet_name: str) -> list[tuple[object, ...]]:
    if sheet_name not in workbook.sheetnames:
        raise PackageValidationError(f"Demo Excel sheet bulunamadı: {sheet_name}")
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if len(rows) < 3:
        raise PackageValidationError(f"Demo Excel sheet boş: {sheet_name}")
    return [cast(tuple[object, ...], row) for row in rows[3:] if row[0] is not None]


def _demo_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise PackageValidationError("Demo Excel tarih hücresi geçersiz.")


def _read_demo_workbooks(root: Path) -> StandardPackage:
    files = sorted(root.glob("*.xlsx"), key=lambda item: item.name)
    if {item.name for item in files} != DEMO_WORKBOOK_NAMES:
        raise PackageValidationError("Demo paketi beklenen beş bina Excelini içermelidir.")
    sites: list[SiteRow] = []
    units: list[ResidentUnitRow] = []
    charges: list[ChargeRow] = []
    payments: list[PaymentRow] = []
    expenses: list[ExpenseRow] = []
    bank_transactions: list[BankTransactionRow] = []
    announcements: list[DemoAnnouncementRow] = []
    file_hashes: list[tuple[str, str]] = []
    istanbul = ZoneInfo("Europe/Istanbul")
    announcement_specs = (
        (
            "2026-06",
            "Haziran 2026 Aidat Son Ödeme Tarihi",
            "Haziran 2026 aidatının son ödeme tarihi 20 Haziran 2026’dır.",
            datetime(2026, 6, 5, 9, tzinfo=istanbul),
        ),
        (
            "2026-07",
            "Temmuz 2026 Aidat Son Ödeme Tarihi",
            "Temmuz 2026 aidatının son ödeme tarihi 20 Temmuz 2026’dır.",
            datetime(2026, 7, 5, 9, tzinfo=istanbul),
        ),
        (
            "2026-08",
            "Ağustos 2026 Aidat Son Ödeme Tarihi",
            "Ağustos 2026 aidatının son ödeme tarihi 20 Ağustos 2026’dır.",
            datetime(2026, 8, 5, 9, tzinfo=istanbul),
        ),
    )
    for file_path in files:
        file_hashes.append((file_path.name, _sha256(file_path)))
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except (OSError, ValueError) as error:
            raise PackageValidationError(f"Demo Excel okunamadı: {file_path.name}") from error
        summary = {
            str(row[0]): row[1]
            for row in workbook["Özet"].iter_rows(min_row=3, values_only=True)
            if row[0] is not None
        }
        site_key = str(summary.get("Bina Kodu", "")).strip()
        site_name = str(summary.get("Bina", "")).strip()
        if not site_key or not site_name:
            raise PackageValidationError("Demo Excel bina kimliği eksik.")
        district = str(summary.get("Konum") or "İstanbul")
        site_payload: dict[str, object] = {
            "source_site_key": site_key,
            "site_name": site_name,
            "district": district,
        }
        sites.append(
            SiteRow(
                source_site_key=site_key,
                site_name=site_name,
                site_slug=site_key.lower(),
                city="İstanbul",
                district=district,
                address_line=district,
                is_active=True,
                payload_hash=_payload_hash(site_payload),
            )
        )
        resident_codes: list[str] = []
        for row in _demo_rows(workbook, "Sakinler"):
            unit_number = str(row[0])
            resident_code = str(row[1]).strip()
            resident_codes.append(resident_code)
            data = {
                "site": site_key,
                "unit": unit_number,
                "resident": resident_code,
                "name": str(row[2]).strip(),
                "active": str(row[5]).strip() == "Aktif",
            }
            units.append(
                ResidentUnitRow(
                    source_site_key=site_key,
                    source_unit_key=resident_code,
                    block_name=None,
                    unit_number=unit_number,
                    floor_label=None,
                    resident_source_key=resident_code,
                    resident_full_name=str(row[2]).strip(),
                    phone=None,
                    email=f"{resident_code.lower()}@example.com",
                    is_active=data["active"] is True,
                    initial_password="YapibinaDemo2026!",
                    payload_hash=_payload_hash(data),
                )
            )
        for row in _demo_rows(workbook, "Giderler"):
            contributions = tuple(
                (resident_code, _money(row[7 + index], allow_zero=True))
                for index, resident_code in enumerate(resident_codes)
            )
            amount = _money(row[5])
            contribution_total = sum(
                (value for _, value in contributions), Decimal("0")
            )
            if abs(contribution_total - amount) > Decimal("0.05"):
                raise PackageValidationError("Demo gider payları toplam giderle eşleşmiyor.")
            data = {
                "site": site_key,
                "key": row[6],
                "date": _demo_date(row[0]),
                "amount": amount,
                "contributions": tuple(
                    (key, str(value)) for key, value in contributions
                ),
            }
            expenses.append(
                ExpenseRow(
                    source_site_key=site_key,
                    source_expense_key=str(row[6]),
                    expense_date=_demo_date(row[0]),
                    expense_month=_demo_date(row[1]),
                    category=str(row[2]).strip(),
                    description=str(row[3]).strip(),
                    payment_method=str(row[4]).strip(),
                    amount=amount,
                    contributions=contributions,
                    payload_hash=_payload_hash(data),
                )
            )
        for row in _demo_rows(workbook, "Aidat Ekstresi"):
            period = _demo_date(row[0])
            resident_code = str(row[2]).strip()
            amount = _money(row[16])
            paid = _money(row[17], allow_zero=True)
            remaining = _money(row[19], allow_zero=True)
            if abs(amount - paid - remaining) > Decimal("0.02"):
                raise PackageValidationError("Demo aidat toplamı ödeme ve kalanla eşleşmiyor.")
            charge_data = {
                "site": site_key, "resident": resident_code, "period": period, "amount": amount
            }
            charges.append(
                ChargeRow(
                    source_site_key=site_key,
                    source_charge_key=f"{resident_code}-AIDAT-{period:%Y-%m}",
                    source_unit_key=resident_code,
                    charge_date=period,
                    due_date=date(period.year, period.month, 20),
                    period_year=period.year,
                    period_month=period.month,
                    charge_type="monthly_due",
                    title=f"{period:%m/%Y} Aidatı",
                    description=f"{_demo_date(row[1]):%m/%Y} ortak gider payı",
                    amount=amount,
                    payload_hash=_payload_hash(charge_data),
                )
            )
            if paid > 0:
                payment_date = _demo_date(row[18])
                payment_data = {
                    "site": site_key, "resident": resident_code, "period": period, "amount": paid
                }
                payments.append(
                    PaymentRow(
                        source_site_key=site_key,
                        source_payment_key=f"{resident_code}-ODEME-{period:%Y-%m}",
                        source_unit_key=resident_code,
                        payment_date=payment_date,
                        payment_method="bank_transfer",
                        amount=paid,
                        reference=f"{resident_code}-{period:%m}",
                        description=f"{period:%m/%Y} aidat ödemesi",
                        target_charge_source_key=(
                            f"{resident_code}-AIDAT-{period:%Y-%m}"
                        ),
                        payload_hash=_payload_hash(payment_data),
                    )
                )
        previous_balance: Decimal | None = None
        for row_number, row in enumerate(_demo_rows(workbook, "Banka Hareketleri"), 1):
            inflow = _money(row[3], allow_zero=True)
            outflow = _money(row[4], allow_zero=True)
            balance = _money(row[5], allow_zero=True)
            if inflow > 0 and outflow > 0:
                raise PackageValidationError("Banka hareketi aynı anda giriş ve çıkış olamaz.")
            balance_difference = (
                previous_balance + inflow - outflow - balance
                if previous_balance is not None
                else Decimal("0")
            )
            if previous_balance is not None and abs(balance_difference) > Decimal("0.02"):
                raise PackageValidationError("Banka hareketi bakiyesi tutarsız.")
            previous_balance = balance
            reference = str(row[7] or f"{site_key}-BANK-{row_number:04d}")
            data = {
                "site": site_key,
                "reference": reference,
                "date": _demo_date(row[0]),
                "balance": balance,
            }
            bank_transactions.append(
                BankTransactionRow(
                    source_site_key=site_key,
                    source_transaction_key=f"{reference}-{row_number:04d}",
                    transaction_date=_demo_date(row[0]),
                    description=str(row[1]).strip(),
                    transaction_type=str(row[2]).strip(),
                    inflow=inflow,
                    outflow=outflow,
                    balance=balance,
                    category=str(row[6]).strip(),
                    reference=reference,
                    payload_hash=_payload_hash(data),
                )
            )
        for suffix, title, body, published_at in announcement_specs:
            key = f"{site_key}-DUYURU-{suffix}"
            announcements.append(
                DemoAnnouncementRow(
                    source_site_key=site_key,
                    source_announcement_key=key,
                    title=title,
                    body=body,
                    published_at=published_at,
                    payload_hash=_payload_hash(
                        {
                            "key": key,
                            "title": title,
                            "body": body,
                            "published_at": published_at,
                        }
                    ),
                )
            )
    fingerprint_material = json.dumps(
        file_hashes, ensure_ascii=False, separators=(",", ":")
    ).encode()
    fingerprint = hashlib.sha256(fingerprint_material).hexdigest()
    return StandardPackage(
        root=root,
        dataset_name="Yapıbina Beş Bina Demo",
        dataset_version="2026.07",
        schema_version="demo-building-v1",
        manifest_sha256=fingerprint,
        fingerprint=fingerprint,
        sites=tuple(sites),
        units=tuple(units),
        charges=tuple(charges),
        payments=tuple(payments),
        expense_count=len(expenses),
        announcement_count=len(announcements),
        expenses=tuple(expenses),
        bank_transactions=tuple(bank_transactions),
        demo_announcements=tuple(announcements),
    )


def validate_package_relationships(package: StandardPackage) -> None:
    site_keys = [row.source_site_key for row in package.sites]
    unit_keys = [row.source_unit_key for row in package.units]
    resident_keys = [row.resident_source_key for row in package.units]
    charge_keys = [row.source_charge_key for row in package.charges]
    payment_keys = [row.source_payment_key for row in package.payments]
    expense_keys = [row.source_expense_key for row in package.expenses]
    bank_keys = [row.source_transaction_key for row in package.bank_transactions]
    announcement_keys = [
        row.source_announcement_key for row in package.demo_announcements
    ]
    for values, label in (
        (site_keys, "source_site_key"),
        (unit_keys, "source_unit_key"),
        (resident_keys, "resident_source_key"),
        (charge_keys, "source_charge_key"),
        (payment_keys, "source_payment_key"),
        (expense_keys, "source_expense_key"),
        (bank_keys, "source_transaction_key"),
        (announcement_keys, "source_announcement_key"),
    ):
        if len(values) != len(set(values)):
            raise PackageValidationError(f"{label} benzersiz olmalıdır.")
    site_set = set(site_keys)
    unit_set = set(unit_keys)
    if any(row.source_site_key not in site_set for row in package.units):
        raise PackageValidationError("Unit site referansı geçersiz.")
    charge_reference_invalid = any(
        row.source_site_key not in site_set or row.source_unit_key not in unit_set
        for row in package.charges
    )
    payment_reference_invalid = any(
        row.source_site_key not in site_set or row.source_unit_key not in unit_set
        for row in package.payments
    )
    if charge_reference_invalid or payment_reference_invalid:
        raise PackageValidationError("Finans kaydı unit/site referansı geçersiz.")
    if any(
        row.source_site_key not in site_set
        or any(unit_key not in unit_set for unit_key, _ in row.contributions)
        for row in package.expenses
    ):
        raise PackageValidationError("Gider contribution referansı geçersiz.")

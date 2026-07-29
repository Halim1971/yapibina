from __future__ import annotations

import json
import shutil
import uuid
from dataclasses import replace
from pathlib import Path

import pytest
from flask import Flask
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import func, select

from app.extensions import db
from app.imports.constants import SOURCE_SYSTEM_STANDARD_EXCEL
from app.imports.exceptions import (
    ConcurrentImportError,
    CriticalFinancialChangeError,
    PackageValidationError,
)
from app.imports.models import ExternalRecordMap, ImportRun, ImportRunStatus
from app.imports.reader import read_standard_package
from app.imports.service import import_standard_package
from app.models import (
    Apartment,
    ApartmentMembership,
    Building,
    Charge,
    Organization,
    OrganizationStatus,
    Payment,
    PaymentAllocation,
)
from app.services import EntityNotFoundError
from app.services.payments import get_payment_unallocated_amount
from scripts.demo_data_lib import canonicalize_xlsx, sha256_file

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEMO_DATA = PROJECT_ROOT / "demo_data"


@pytest.fixture(autouse=True)
def _application_context(app: Flask) -> None:
    del app


def _organization(slug: str = "import-target") -> Organization:
    organization = Organization(
        name="Importer Hedefi",
        slug=slug,
        status=OrganizationStatus.ACTIVE,
    )
    db.session.add(organization)
    db.session.commit()
    return organization


def _copy_package(tmp_path: Path) -> Path:
    target = tmp_path / "package"
    shutil.copytree(DEMO_DATA, target)
    return target


def _update_manifest(package_path: Path, relative_path: str) -> None:
    manifest_path = package_path / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    workbook_path = package_path / relative_path
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    row_count = workbook.active.max_row - 1
    for item in manifest["files"]:
        if item["relative_path"] == relative_path:
            item["row_count"] = row_count
            item["sha256"] = sha256_file(workbook_path)
            break
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _change_cell(
    package_path: Path,
    relative_path: str,
    *,
    source_key_column: str,
    source_key: str,
    column: str,
    value: object,
) -> None:
    workbook_path = package_path / relative_path
    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    headers = {
        str(cell.value): index
        for index, cell in enumerate(worksheet[1], start=1)
    }
    for row_index in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_index, headers[source_key_column]).value == source_key:
            worksheet.cell(row_index, headers[column]).value = value
            break
    workbook.save(workbook_path)
    canonicalize_xlsx(workbook_path)
    _update_manifest(package_path, relative_path)


def test_reader_accepts_package_and_rejects_manifest_or_schema(
    tmp_path: Path,
) -> None:
    package = read_standard_package(DEMO_DATA)
    assert len(package.sites) == 5
    assert len(package.units) == 50
    assert len(package.charges) == 305
    assert len(package.payments) == 255

    corrupted = _copy_package(tmp_path)
    (corrupted / "sites.xlsx").write_bytes(b"not an Excel file")
    with pytest.raises(PackageValidationError, match="hash"):
        read_standard_package(corrupted)

    invalid_schema = _copy_package(tmp_path / "schema")
    manifest_path = invalid_schema / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["schema_version"] = "99.0"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    with pytest.raises(PackageValidationError, match="schema"):
        read_standard_package(invalid_schema)


def test_full_import_is_tenant_safe_allocated_and_idempotent() -> None:
    organization = _organization()
    package = read_standard_package(DEMO_DATA)
    result = import_standard_package(
        db.session,
        organization_id=organization.id,
        package=package,
        source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
    )
    assert result.status == "completed"
    assert result.inserted == 665
    assert result.deferred == 155
    assert db.session.scalar(
        select(func.count()).select_from(Building).where(
            Building.organization_id == organization.id
        )
    ) == 5
    assert db.session.scalar(
        select(func.count()).select_from(Apartment).where(
            Apartment.organization_id == organization.id
        )
    ) == 50
    assert db.session.scalar(
        select(func.count()).select_from(ApartmentMembership).where(
            ApartmentMembership.organization_id == organization.id,
            ApartmentMembership.is_active.is_(True),
        )
    ) == 50
    assert db.session.scalar(
        select(func.count()).select_from(Charge).where(
            Charge.organization_id == organization.id
        )
    ) == 305
    assert db.session.scalar(
        select(func.count()).select_from(Payment).where(
            Payment.organization_id == organization.id
        )
    ) == 255
    assert db.session.scalar(
        select(func.count()).select_from(ExternalRecordMap).where(
            ExternalRecordMap.organization_id == organization.id,
            ExternalRecordMap.entity_type == "resident",
        )
    ) == 50

    payment = db.session.scalar(
        select(Payment)
        .where(Payment.organization_id == organization.id)
        .order_by(Payment.payment_date.desc())
    )
    assert payment is not None
    allocations = db.session.scalars(
        select(PaymentAllocation).where(
            PaymentAllocation.organization_id == organization.id,
            PaymentAllocation.payment_id == payment.id,
        )
    ).all()
    allocated_charges = [
        db.session.get(Charge, allocation.charge_id) for allocation in allocations
    ]
    due_dates = [item.due_date for item in allocated_charges if item is not None]
    assert due_dates == sorted(due_dates)
    assert any(
        get_payment_unallocated_amount(
            db.session,
            organization_id=organization.id,
            payment_id=item.id,
        )
        > 0
        for item in db.session.scalars(
            select(Payment).where(Payment.organization_id == organization.id)
        )
    )

    second = import_standard_package(
        db.session,
        organization_id=organization.id,
        package=package,
        source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
    )
    assert second.status == "already_imported"
    assert second.inserted == 0
    assert second.skipped == 665
    assert db.session.scalar(select(func.count()).select_from(ImportRun)) == 1


def test_import_rejects_missing_target_and_dry_run_is_not_persistent() -> None:
    package = read_standard_package(DEMO_DATA)
    with pytest.raises(EntityNotFoundError):
        import_standard_package(
            db.session,
            organization_id=uuid.uuid4(),
            package=package,
            source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        )
    organization = _organization()
    result = import_standard_package(
        db.session,
        organization_id=organization.id,
        package=package,
        source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        dry_run=True,
    )
    assert result.status == "dry_run"
    assert result.inserted == 665
    assert db.session.scalar(select(func.count()).select_from(Building)) == 0
    assert db.session.scalar(select(func.count()).select_from(ImportRun)) == 0


def test_invalid_relationship_rolls_back_and_marks_run_failed() -> None:
    organization = _organization()
    package = read_standard_package(DEMO_DATA)
    invalid_charge = replace(package.charges[0], source_unit_key="missing-unit")
    invalid_package = replace(
        package,
        fingerprint="f" * 64,
        charges=(invalid_charge, *package.charges[1:]),
    )

    with pytest.raises(PackageValidationError, match="unit"):
        import_standard_package(
            db.session,
            organization_id=organization.id,
            package=invalid_package,
            source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        )
    assert db.session.scalar(select(func.count()).select_from(Building)) == 0
    run = db.session.scalar(select(ImportRun))
    assert run is not None
    assert run.status is ImportRunStatus.FAILED


def test_running_import_guard_rejects_concurrent_run() -> None:
    organization = _organization()
    package = read_standard_package(DEMO_DATA)
    db.session.add(
        ImportRun(
            organization_id=organization.id,
            source_system="another_source",
            dataset_name="guard",
            dataset_version="1",
            schema_version="1",
            manifest_sha256="0" * 64,
            package_fingerprint="1" * 64,
            status=ImportRunStatus.RUNNING,
        )
    )
    db.session.commit()
    with pytest.raises(ConcurrentImportError):
        import_standard_package(
            db.session,
            organization_id=organization.id,
            package=package,
            source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        )


def test_safe_updates_and_critical_financial_changes(
    tmp_path: Path,
) -> None:
    organization = _organization()
    original = read_standard_package(DEMO_DATA)
    import_standard_package(
        db.session,
        organization_id=organization.id,
        package=original,
        source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
    )

    changed = _copy_package(tmp_path)
    _change_cell(
        changed,
        "sites.xlsx",
        source_key_column="source_site_key",
        source_key="SITE-001",
        column="site_name",
        value="Ulubatlı Demo Sitesi",
    )
    _change_cell(
        changed,
        "ulubatli-sitesi/residents_and_units.xlsx",
        source_key_column="resident_source_key",
        source_key="RES-DEMO-001",
        column="phone",
        value="DEMO-0500-999-00-01",
    )
    updated = import_standard_package(
        db.session,
        organization_id=organization.id,
        package=read_standard_package(changed),
        source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
    )
    assert updated.updated == 2
    assert updated.inserted == 0

    critical = _copy_package(tmp_path / "critical")
    _change_cell(
        critical,
        "ulubatli-sitesi/charges.xlsx",
        source_key_column="source_charge_key",
        source_key="SITE-001-CHG-01-2026-02",
        column="amount",
        value=9999,
    )
    with pytest.raises(CriticalFinancialChangeError):
        import_standard_package(
            db.session,
            organization_id=organization.id,
            package=read_standard_package(critical),
            source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        )
    failed = db.session.scalar(
        select(ImportRun).where(ImportRun.status == ImportRunStatus.FAILED)
    )
    assert failed is not None
    unchanged = db.session.scalar(
        select(Charge)
        .join(
            ExternalRecordMap,
            ExternalRecordMap.internal_id == Charge.id,
        )
        .where(ExternalRecordMap.source_key == "SITE-001-CHG-01-2026-02")
    )
    assert unchanged is not None
    assert str(unchanged.original_amount) != "9999.00"

    critical_payment = _copy_package(tmp_path / "critical-payment")
    _change_cell(
        critical_payment,
        "ulubatli-sitesi/payments.xlsx",
        source_key_column="source_payment_key",
        source_key="SITE-001-PAY-01-01",
        column="amount",
        value=8888,
    )
    with pytest.raises(CriticalFinancialChangeError):
        import_standard_package(
            db.session,
            organization_id=organization.id,
            package=read_standard_package(critical_payment),
            source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        )


def test_same_source_keys_are_isolated_between_organizations() -> None:
    package = read_standard_package(DEMO_DATA)
    first = _organization("first-import")
    second = _organization("second-import")
    for organization in (first, second):
        result = import_standard_package(
            db.session,
            organization_id=organization.id,
            package=package,
            source_system=SOURCE_SYSTEM_STANDARD_EXCEL,
        )
        assert result.status == "completed"
    assert db.session.scalar(select(func.count()).select_from(Building)) == 10
    assert db.session.scalar(
        select(func.count()).select_from(ExternalRecordMap).where(
            ExternalRecordMap.source_key == "SITE-001",
            ExternalRecordMap.entity_type == "site",
        )
    ) == 2


def test_cli_reports_success_and_failure(app: Flask) -> None:
    organization = _organization()
    runner = app.test_cli_runner()
    success = runner.invoke(
        args=[
            "import-standard-data",
            "--organization-id",
            str(organization.id),
            "--path",
            str(DEMO_DATA),
            "--dry-run",
        ]
    )
    assert success.exit_code == 0
    assert "dry_run" in success.output
    failure = runner.invoke(
        args=[
            "import-standard-data",
            "--organization-id",
            str(uuid.uuid4()),
            "--path",
            str(DEMO_DATA),
        ]
    )
    assert failure.exit_code != 0

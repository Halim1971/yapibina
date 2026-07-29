from __future__ import annotations

import json
import shutil
import subprocess
import sys
from collections.abc import Iterator
from decimal import Decimal
from pathlib import Path
from typing import cast

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from scripts.demo_data_lib import (
    MONTHS,
    SITE_SLUGS,
    canonicalize_xlsx,
    read_workbook,
    sha256_file,
)
from scripts.generate_demo_data import generate_demo_data
from scripts.validate_demo_data import DemoValidationError, validate_demo_data

PROJECT_ROOT = Path(__file__).resolve().parents[2]


@pytest.fixture(scope="module")
def generated_data(tmp_path_factory: pytest.TempPathFactory) -> Iterator[Path]:
    output = tmp_path_factory.mktemp("demo-dataset") / "demo_data"
    generate_demo_data(output)
    yield output


def _manifest(path: Path) -> dict[str, object]:
    return cast(
        dict[str, object],
        json.loads((path / "manifest.json").read_text(encoding="utf-8")),
    )


def _file_entry(manifest: dict[str, object], relative_path: str) -> dict[str, object]:
    files = manifest["files"]
    assert isinstance(files, list)
    return next(item for item in files if item["relative_path"] == relative_path)


def test_generator_and_validator_accept_complete_dataset(generated_data: Path) -> None:
    validate_demo_data(generated_data)
    manifest = _manifest(generated_data)
    assert manifest["site_count"] == 5
    assert manifest["unit_count"] == 50
    assert manifest["resident_count"] == 50


def test_each_site_has_ten_unique_units_and_total_fifty_residents(
    generated_data: Path,
) -> None:
    all_residents: set[str] = set()
    for slug in SITE_SLUGS:
        rows = read_workbook(generated_data / slug / "residents_and_units.xlsx")
        assert len(rows) == 10
        unit_keys = [row["source_unit_key"] for row in rows]
        assert len(unit_keys) == len(set(unit_keys))
        all_residents.update(str(row["resident_source_key"]) for row in rows)
    assert len(all_residents) == 50


def test_site_and_source_keys_are_unique(generated_data: Path) -> None:
    sites = read_workbook(generated_data / "sites.xlsx")
    assert len({row["source_site_key"] for row in sites}) == 5
    assert len({row["site_slug"] for row in sites}) == 5

    charge_keys: set[str] = set()
    payment_keys: set[str] = set()
    for slug in SITE_SLUGS:
        for filename, field, collected in (
            ("charges.xlsx", "source_charge_key", charge_keys),
            ("payments.xlsx", "source_payment_key", payment_keys),
        ):
            rows = read_workbook(generated_data / slug / filename)
            new_keys = {str(row[field]) for row in rows}
            assert not collected.intersection(new_keys)
            collected.update(new_keys)


def test_charge_and_payment_foreign_keys_are_valid(generated_data: Path) -> None:
    for slug in SITE_SLUGS:
        site_path = generated_data / slug
        units = {
            row["source_unit_key"]
            for row in read_workbook(site_path / "residents_and_units.xlsx")
        }
        for filename in ("charges.xlsx", "payments.xlsx"):
            assert all(
                row["source_unit_key"] in units
                for row in read_workbook(site_path / filename)
            )


def test_every_unit_has_six_months_of_monthly_charges(
    generated_data: Path,
) -> None:
    for slug in SITE_SLUGS:
        charges = read_workbook(generated_data / slug / "charges.xlsx")
        unit_keys = {str(row["source_unit_key"]) for row in charges}
        for unit_key in unit_keys:
            periods = {
                int(str(row["period_month"]))
                for row in charges
                if row["source_unit_key"] == unit_key
                and row["charge_type"] == "monthly_due"
            }
            assert periods == set(MONTHS)


def test_controlled_partial_overpayment_bulk_and_additional_scenarios(
    generated_data: Path,
) -> None:
    for slug in SITE_SLUGS:
        residents = read_workbook(generated_data / slug / "residents_and_units.xlsx")
        units = [str(row["source_unit_key"]) for row in residents]
        charges = read_workbook(generated_data / slug / "charges.xlsx")
        payments = read_workbook(generated_data / slug / "payments.xlsx")
        charge_totals = {
            unit: sum(
                (Decimal(str(row["amount"])) for row in charges if row["source_unit_key"] == unit),
                start=Decimal("0.00"),
            )
            for unit in units
        }
        payment_totals = {
            unit: sum(
                (Decimal(str(row["amount"])) for row in payments if row["source_unit_key"] == unit),
                start=Decimal("0.00"),
            )
            for unit in units
        }
        monthly = Decimal(str(charges[0]["amount"]))
        assert charge_totals[units[2]] - payment_totals[units[2]] == monthly / 2
        assert payment_totals[units[4]] > charge_totals[units[4]]
        assert any(
            row["source_unit_key"] == units[8]
            and row["charge_type"] == "additional_due"
            for row in charges
        )
        assert max(
            Decimal(str(row["amount"]))
            for row in payments
            if row["source_unit_key"] == units[7]
        ) >= monthly * 3


def test_expense_categories_and_announcement_priorities_are_valid(
    generated_data: Path,
) -> None:
    expense_categories = {
        "cleaning",
        "electricity",
        "water",
        "elevator_maintenance",
        "security",
        "landscaping",
        "management_service",
        "technical_maintenance",
        "insurance",
        "repair",
    }
    for slug in SITE_SLUGS:
        expenses = read_workbook(generated_data / slug / "expenses.xlsx")
        announcements = read_workbook(generated_data / slug / "announcements.xlsx")
        assert len(expenses) == 24
        assert all(row["category"] in expense_categories for row in expenses)
        assert len(announcements) == 7
        assert all(
            row["priority"] in {"normal", "important", "urgent"}
            for row in announcements
        )


def test_all_emails_and_phones_are_explicitly_demo_data(
    generated_data: Path,
) -> None:
    for slug in SITE_SLUGS:
        residents = read_workbook(generated_data / slug / "residents_and_units.xlsx")
        assert all(str(row["email"]).endswith("@example.com") for row in residents)
        assert all(str(row["phone"]).startswith("DEMO-") for row in residents)


def test_manifest_row_counts_and_hashes_are_correct(generated_data: Path) -> None:
    manifest = _manifest(generated_data)
    files = manifest["files"]
    assert isinstance(files, list)
    assert len(files) == 26
    for item in files:
        path = generated_data / str(item["relative_path"])
        assert len(read_workbook(path)) == item["row_count"]
        assert sha256_file(path) == item["sha256"]


def test_generation_is_byte_deterministic(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"
    generate_demo_data(first)
    generate_demo_data(second)
    first_files = {
        path.relative_to(first): path.read_bytes()
        for path in first.rglob("*")
        if path.is_file()
    }
    second_files = {
        path.relative_to(second): path.read_bytes()
        for path in second.rglob("*")
        if path.is_file()
    }
    assert first_files == second_files


def test_excel_files_have_readable_non_decorative_format(generated_data: Path) -> None:
    workbook = load_workbook(generated_data / "ulubatli-sitesi" / "charges.xlsx")
    assert len(workbook.sheetnames) == 1
    worksheet = workbook.active
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions
    assert all(cell.font.bold for cell in worksheet[1])
    assert not worksheet.merged_cells.ranges
    assert worksheet.sheet_state == "visible"
    assert all(
        not (isinstance(cell.value, str) and cell.value.startswith("="))
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_validator_rejects_broken_foreign_key_even_with_updated_hash(
    generated_data: Path,
    tmp_path: Path,
) -> None:
    broken = tmp_path / "broken-foreign-key"
    shutil.copytree(generated_data, broken)
    relative = "ulubatli-sitesi/charges.xlsx"
    charge_path = broken / relative
    workbook = load_workbook(charge_path)
    worksheet = workbook.active
    headers = {cell.value: cell.column for cell in worksheet[1]}
    worksheet.cell(row=2, column=headers["source_unit_key"], value="UNKNOWN-UNIT")
    workbook.save(charge_path)
    canonicalize_xlsx(charge_path)
    manifest = _manifest(broken)
    _file_entry(manifest, relative)["sha256"] = sha256_file(charge_path)
    (broken / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    with pytest.raises(DemoValidationError, match="charge unit referansı geçersiz"):
        validate_demo_data(broken, check_determinism=False)


def test_validator_rejects_modified_file_hash(
    generated_data: Path,
    tmp_path: Path,
) -> None:
    broken = tmp_path / "broken-hash"
    shutil.copytree(generated_data, broken)
    path = broken / "ulubatli-sitesi" / "payments.xlsx"
    path.write_bytes(path.read_bytes() + b"changed")
    with pytest.raises(DemoValidationError, match="SHA-256"):
        validate_demo_data(broken, check_determinism=False)


def test_validator_cli_returns_nonzero_for_invalid_dataset(tmp_path: Path) -> None:
    missing = tmp_path / "missing"
    result = subprocess.run(
        [
            sys.executable,
            "scripts/validate_demo_data.py",
            "--path",
            str(missing),
        ],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode != 0
    assert "doğrulaması başarısız" in result.stdout
